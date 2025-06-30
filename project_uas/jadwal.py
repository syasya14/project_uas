import pandas as pd
from datetime import datetime, timedelta, time
from collections import defaultdict
import re
import openpyxl
from openpyxl.styles import PatternFill

print("\n🗕️ Menjalankan penjadwalan utama...")

# ================== Konfigurasi Awal ==================
AVAILABLE_ROOMS = {
    'GD A': {
        2: [f"A2-{i}" for i in range(1, 9)],
        3: [f"A3-{i}" for i in range(1, 9)],
        4: [f"A4-{i}" for i in range(1, 9)],
        5: [f"A5-{i}" for i in range(1, 9)],
    },
    'GD B': {
        3: [f"B3-{i}" for i in range(1, 6)],
        4: [f"B4-{i}" for i in range(1, 6)],
        5: [f"B5-{i}" for i in range(1, 6)],
    }
}

ROOM_PREFERENCES = {
    'TI': {'floors': [3, 4]},
    'SI': {'floors': [3, 4]},
    'DK': {'floors': [4, 5]},
    'SD': {'floors': [2, 3]},
    'HK': {'floors': [3, 4]},
    'ME': {'floors': [4, 5]},
    'EL': {'floors': [4, 5]},
    'AKT': {'floors': [2, 3]},
    'MJN': {'floors': [2, 3]},
}

# ================== Baca dan Persiapkan Data ==================
df = pd.read_excel("Data_pengajaran.xlsx", sheet_name="Sheet1")
df["Available Day"] = df["Available Day"].fillna("ALL").str.upper()
df["Available Times"] = df["Available Times"].fillna("ALL").astype(str).str.upper()
df["Kelas"] = df["Kelas"].str.upper()

DAYS = ["SENIN", "SELASA", "RABU", "KAMIS", "JUMAT", "SABTU", "MINGGU"]
ISTIRAHAT_LIST = [(time(12, 0), time(13, 0)), (time(18, 0), time(18, 30))]

def sks_to_duration(sks): return timedelta(minutes=50 * sks)
def is_malam(kelas): return "M" in kelas
def get_allowed_days(kelas):
    if "B" in kelas: return ["SABTU"]
    elif "C" in kelas: return ["MINGGU"]
    else: return DAYS

def get_time_window(kelas):
    if is_malam(kelas):
        return (time(17, 0), time(22, 0))
    elif "B" in kelas or "C" in kelas:
        return (time(8, 0), time(21, 0))
    else:
        return (time(8, 0), time(18, 0))

def generate_slots(start, end, step=10):
    slots, t = [], datetime.combine(datetime.today(), start)
    while t.time() < end:
        slots.append(t.time())
        t += timedelta(minutes=step)
    return slots

def is_in_istirahat(start, end):
    return any(start < i_end and end > i_start for i_start, i_end in ISTIRAHAT_LIST)

def is_conflict(jadwal, hari, entitas, start, end):
    return any(not (end <= s or start >= e) for s, e in jadwal[hari][entitas])

def cari_ruangan(hari, start, end, kelas, jadwal):
    prodi = re.match(r"([A-Z]{2})", kelas)
    if not prodi:
        return None
    kode = prodi.group(1)
    preferensi = ROOM_PREFERENCES.get(kode, {"floors": []})["floors"]
    for gedung, lantai_dict in AVAILABLE_ROOMS.items():
        for lantai in preferensi:
            for ruang in lantai_dict.get(lantai, []):
                if not is_conflict(jadwal, hari, ruang, start, end):
                    return ruang
    return None

def cari_slot(hari, kelas, dosen, durasi, allowed_times, jadwal):
    start_win, end_win = get_time_window(kelas)
    for start in generate_slots(start_win, end_win):
        end = (datetime.combine(datetime.today(), start) + durasi).time()
        if end > end_win or is_in_istirahat(start, end): continue

        if allowed_times != "ALL":
            try:
                a_start_str = allowed_times.split("-")[0].strip()
                if ":" not in a_start_str: raise ValueError("Format waktu tidak valid")
                a_start = datetime.strptime(a_start_str, "%H:%M").time()
                if start < a_start: continue
            except Exception as e:
                print(f"[⚠] Gagal parsing Available Times '{allowed_times}' → {e}")
                pass

        if is_conflict(jadwal, hari, dosen, start, end): continue
        if is_conflict(jadwal, hari, kelas, start, end): continue

        return start, end
    return None, None

# ================== Penjadwalan ==================
final_schedule = []
gagal_terjadwal = []
occupied_times = defaultdict(lambda: defaultdict(list))

for _, row in df.iterrows():
    dosen, matkul, sks = row["DOSEN"], row["Mata Kuliah"], int(row["SKS"])
    kelas_list = row["Kelas"].split(",")
    available_days = row["Available Day"].split(",") if row["Available Day"] != "ALL" else DAYS
    available_times = row["Available Times"]
    duration = sks_to_duration(sks)

    for kelas in kelas_list:
        kelas = kelas.strip()
        hari_pilihan = [h for h in available_days if h in get_allowed_days(kelas)]
        scheduled = False
        for hari in hari_pilihan:
            start, end = cari_slot(hari, kelas, dosen, duration, available_times, occupied_times)
            if start and end:
                ruang = cari_ruangan(hari, start, end, kelas, occupied_times)
                if not ruang:
                    continue

                jadwal_kelas_hari_ini = [
                    j for j in final_schedule
                    if j["Kelas"] == kelas and j["Hari"] == hari and j["Status"] in ["TERJADWAL", "ONLINE"]
                ]

                if any(k in kelas for k in ["B", "C", "M"]):
                    if len(jadwal_kelas_hari_ini) >= 10:
                        continue
                else:
                    if len(jadwal_kelas_hari_ini) >= 3:
                        continue

                status = "ONLINE" if end > time(21, 0) else "TERJADWAL"

                final_schedule.append({
                    "Dosen": dosen, "Mata Kuliah": matkul, "Kelas": kelas,
                    "Hari": hari, "Jam": f"{start.strftime('%H:%M')} - {end.strftime('%H:%M')}",
                    "Ruangan": ruang, "Tanggal": pd.NaT, "Status": status
                })
                occupied_times[hari][dosen].append((start, end))
                occupied_times[hari][kelas].append((start, end))
                occupied_times[hari][ruang].append((start, end))
                scheduled = True
                break

        if not scheduled:
            start_win, _ = get_time_window(kelas)
            start_time = start_win
            end_time = (datetime.combine(datetime.today(), start_time) + duration).time()
            final_schedule.append({
                "Dosen": dosen, "Mata Kuliah": matkul, "Kelas": kelas,
                "Hari": "ONLINE", "Jam": f"{start_time.strftime('%H:%M')} - {end_time.strftime('%H:%M')}",
                "Ruangan": "-", "Tanggal": pd.NaT, "Status": "ONLINE"
            })
            gagal_terjadwal.append({
                "Dosen": dosen, "Mata Kuliah": matkul, "Kelas": kelas,
                "Alasan": "Tidak ditemukan slot tersedia sesuai SKS dan waktu dosen",
                "Available Day": ", ".join(available_days),
                "Available Times": available_times,
                "SKS": sks
            })

# ================== Simpan Output ==================
def extract_prodi_angkatan(kelas):
    match = re.match(r"([A-Z]{2})(\\d{2})", str(kelas).upper())
    if match:
        return match.group(1) + "20" + match.group(2)
    return "LAINNYA"

jadwal_df = pd.DataFrame(final_schedule)
jadwal_df["Sheet"] = jadwal_df["Kelas"].apply(extract_prodi_angkatan)

with pd.ExcelWriter("jadwal_output_dengan_pengganti.xlsx", engine="openpyxl") as writer:
    for sheet_name, subdf in jadwal_df.groupby("Sheet"):
        subdf.drop(columns=["Sheet"]).to_excel(writer, sheet_name=sheet_name, index=False)
        sheet = writer.book[sheet_name]
        for row_idx, row in subdf.iterrows():
            if row["Status"] == "ONLINE":
                for col_idx in range(1, len(subdf.columns) + 1):
                    sheet.cell(row=row_idx + 2, column=col_idx).fill = PatternFill("solid", fgColor="FFFF00")

    def clean_sheet_name(name):
        return re.sub(r"[\\/*?:\\[\\]]", "_", name)[:31]

    for dosen_name, subdf in jadwal_df.groupby("Dosen"):
        safe_name = clean_sheet_name(dosen_name)
        subdf.to_excel(writer, sheet_name=safe_name, index=False)

    rekap_matkul = jadwal_df.groupby("Kelas")["Mata Kuliah"].unique().reset_index()
    rekap_matkul["Mata Kuliah"] = rekap_matkul["Mata Kuliah"].apply(
        lambda x: ", ".join(sorted({str(i) for i in x if pd.notna(i)}))
    )
    rekap_matkul.to_excel(writer, sheet_name="REKAP_MATKUL_PER_KELAS", index=False)
    pd.DataFrame(gagal_terjadwal).to_excel(writer, sheet_name="GAGAL_TERJADWAL", index=False)

print("✅ Jadwal berhasil disimpan ke jadwal_output_dengan_pengganti.xlsx")
